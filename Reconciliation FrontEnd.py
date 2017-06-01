import openpyxl
from openpyxl.reader.excel import load_workbook
from tkinter import *
from tkinter import messagebox
from tkinter.filedialog import askopenfilename

def selectFiles():
    Sales2 = Sales.get()
    Stock2 = Stock.get()
    if Sales2 == 0 and Stock2 == 0:
        messagebox.showwarning("Error", "You need to select sales or stock")
    else:
        root.withdraw
        #User inputs for location of files and whether Sales or Stock for reconciliation
        BOReport = askopenfilename(parent=root,initialdir="/",title='Please select a BO Report')
        INCReport = askopenfilename(parent=root,initialdir="/",title='Please select an INC Report')
        Reconciliation(Sales2,BOReport, INCReport)

def Reconciliation(Sales,BOReport, INCReport):

    #Runs Sales Reconciliation
    def SalesRec(BOReport, INCReport):
        selectFiles
        #Sets Excel files to reports
        BO = load_workbook(BOReport)
        INC = load_workbook(INCReport)
        print("See me Sales")

    #Runs Stock Reconciliation
    def StockRec(BOReport, INCReport):
        selectFiles
        #Sets Excel files to reports
        print(BOReport)
        print(INCReport)
        BO = load_workbook(BOReport)
        INC = load_workbook(INCReport)
        print("Watch me stock, stock")

    if Sales == 1:
        SalesRec(BOReport, INCReport)
    else:
        StockRec(BOReport, INCReport)

#Sets Tkinter root window
root = Tk()

Label (root, text = "Sales or Stock").grid(row=0, sticky=W)
Sales = IntVar()
Checkbutton(root, text="Sales", variable = Sales).grid(row=1, sticky=W)
Stock = IntVar()
Checkbutton(root, text="Stock", variable = Stock).grid(row=2, sticky=W)
Button(root, text = 'Quit', command=root.quit).grid(row=4, sticky=W, pady=4)
Button(root, text = 'Next', command = selectFiles).grid(row=3, sticky=W, pady=4)
mainloop()

