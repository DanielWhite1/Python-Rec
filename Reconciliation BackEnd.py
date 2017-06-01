import openpyxl
import FrontEnd

#Checks sales or stock based on the GUI
if FrontEnd.Sales == 1:
    Sales(FrontEnd.BOReport, FrontEnd.INCReport)
else:
    Stock(FrontEnd.BOReport, FrontEnd.INCReport)

#Runs Sales Reconciliation
def Sales(BOReport, INCReport):
    #Sets Excel files to reports
    BO = Workbook(BOReport)
    INC = Workbook(INCReport)
    print("See me Sales")

#Runs Stock Reconciliation
def Stock(BOReport, INCReport):
    #Sets Excel files to reports
    BO = Workbook(BOReport)
    INC = Workbook(INCReport)
    print("Watch me stock, stock")
